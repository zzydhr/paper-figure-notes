"""L4 渲染层 —— 长表（Excel / CSV）。

**一行 = 一个实验组**。这是跨文献横向比较的行粒度：同一个筛选条件
（「所有用了 U251 的实验」「所有 3 μM 剂量组」「所有 seed=3 的 ablation」）
必须能靠一次 Excel 筛选选出来，所以论文级 / 图级 / panel 级的字段全部
冗余地摊平到每一行上。

产物：
    render_tables()       单篇 → groups.xlsx + groups.csv
    render_corpus_table() 多篇 → all_groups.xlsx（+ all_groups.csv）

Windows 编码：xlsx 二进制无所谓，CSV 必须 `utf-8-sig`，否则 Excel
双击打开中文全是乱码。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pfn.models import (
    FigureRecord,
    Panel,
    PaperBundle,
    is_reported,
    needs_review,
)

MISSING = "未说明"

TABLE_BASENAME = "groups"
CORPUS_BASENAME = "all_groups"

SHEET_GROUPS = "实验组长表"
SHEET_PAPERS = "论文清单"
SHEET_SUMMARY = "图速览"

#: 四列速览表在 Excel 里的列序。前面多一列 figure_id 用于分组，
#: 且**不留空重复值**——markdown 靠留空表达「同属一个 panel」，
#: 但 Excel 一排序空单元格就和上面的行脱钩了。
SUMMARY_COLUMNS = ["figure_id", "小图", "实验方法", "分组", "结果"]

#: 长表列序：论文身份 → 图/panel 定位 → 实验设计 → 组本身 → 统计 → 结论 → 溯源。
#: 顺序即阅读顺序，改动会影响已有汇总表的对齐，慎改。
COLUMNS: list[str] = [
    "论文标题",
    "年份",
    "venue",
    "DOI",
    "domain",
    "figure_id",
    "图类型",
    "页码",
    "panel",
    "实验描述",
    "自变量",
    "水平",
    "受控条件",
    "实验单位",
    "组名",
    "组角色",
    "处理条件",
    "n",
    "测量指标",
    "结果",
    "vs对照",
    "统计检验",
    "误差棒",
    "重复",
    "显著性标注",
    "panel结论",
    "整图结论",
    "置信度",
    "是否需复核",
    "推断字段",
    "证据来源",
    "图片",
]

CORPUS_COLUMNS: list[str] = ["来源"] + COLUMNS

PAPER_COLUMNS: list[str] = [
    "来源",
    "论文标题",
    "年份",
    "venue",
    "DOI",
    "domain",
    "图数",
    "panel数",
    "实验组数",
    "需复核panel数",
    "源文件",
]

#: 这些列内容长，Excel 里强制换行；其余列按内容宽度自适应
WRAP_COLUMNS = {
    "实验方法",
    "分组",
    "结果",
    "论文标题",
    "实验描述",
    "自变量",
    "水平",
    "受控条件",
    "处理条件",
    "测量指标",
    "结果",
    "vs对照",
    "统计检验",
    "误差棒",
    "显著性标注",
    "panel结论",
    "整图结论",
    "推断字段",
    "证据来源",
}

MIN_WIDTH, MAX_WIDTH = 8, 42

HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REVIEW_FILL = PatternFill("solid", fgColor="FFE0B2")

EVIDENCE_SOURCES = (
    ("图题", "from_caption"),
    ("正文", "from_body"),
    ("Methods", "from_methods"),
    ("读图", "from_image_only"),
)


# ── 取值 ────────────────────────────────────────────────────────────────────


def _val(value: Any) -> str:
    return str(getattr(value, "value", value) or "")


def _text(value: Any, default: str = MISSING) -> str:
    """not_reported → 「未说明」。留空会让「工具没填」和「原文没写」混为一谈。"""
    if not is_reported(value):
        return default
    return str(value).strip()


def _join(items: Any, sep: str = "；", default: str = MISSING) -> str:
    if not is_reported(items):
        return default
    parts = [str(i).strip() for i in items if str(i).strip()]
    return sep.join(parts) if parts else default


def _evidence_summary(panel: Panel) -> str:
    parts = []
    for label, attr in EVIDENCE_SOURCES:
        values = getattr(panel.evidence, attr)
        if values:
            parts.append(f"{label}: {', '.join(values)}")
    return "；".join(parts) if parts else MISSING


def _dw(text: str) -> int:
    """显示宽度：CJK 按 2 算，近似 Excel 的列宽单位。"""
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in text)


# ── 行构造 ──────────────────────────────────────────────────────────────────


def _figure_rows(bundle: PaperBundle, fig: FigureRecord) -> list[dict[str, Any]]:
    meta = fig.paper if fig.paper.title.strip() else bundle.paper
    common = {
        "论文标题": _text(meta.title, ""),
        "年份": meta.year if meta.year is not None else "",
        "venue": _text(meta.venue, ""),
        "DOI": _text(meta.doi, ""),
        "domain": _val(fig.domain) or _val(bundle.domain),
        "figure_id": fig.figure_id,
        "图类型": _val(fig.figure_kind),
        "页码": ", ".join(str(p) for p in fig.pages),
        "整图结论": _text(fig.figure_conclusion),
        "置信度": _val(fig.confidence),
        "图片": "; ".join(img.replace("\\", "/") for img in fig.images),
    }

    rows: list[dict[str, Any]] = []
    for panel in fig.panels:
        panel_common = {
            **common,
            "panel": panel.label or "-",
            "实验描述": _text(panel.experiment),
            "自变量": _text(panel.design.independent_var),
            "水平": _join(panel.design.levels, " / "),
            "受控条件": _join(panel.design.controlled),
            "实验单位": _text(panel.design.unit),
            "统计检验": _text(panel.stats.test),
            "误差棒": _text(panel.stats.error_bar),
            "重复": _text(panel.stats.replicates),
            "显著性标注": _text(panel.stats.significance),
            "panel结论": _text(panel.finding),
            "是否需复核": "是" if needs_review(panel) else "否",
            "推断字段": _join(panel.evidence.inferred, ", ", ""),
            "证据来源": _evidence_summary(panel),
        }
        # 一行一个实验组：无分组的 panel（示意图 / 定性图）不产生行
        for group in panel.groups:
            rows.append(
                {
                    **panel_common,
                    "组名": _text(group.name, "-"),
                    "组角色": _val(group.role),
                    "处理条件": _text(group.condition),
                    "n": _text(group.n),
                    "测量指标": _text(group.readout),
                    "结果": _text(group.result),
                    "vs对照": _text(group.vs_control),
                }
            )
    return rows


def _bundle_rows(bundle: PaperBundle, source: str | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for fig in bundle.figures:
        rows.extend(_figure_rows(bundle, fig))
    if source is not None:
        for row in rows:
            row["来源"] = source
    return rows


def _bundle_source(bundle: PaperBundle) -> str:
    """汇总表的第一列：优先输出目录名，退回 PDF 名，再退回标题。"""
    for candidate in (bundle.output_dir, bundle.source_pdf):
        if candidate.strip():
            return Path(candidate).name
    return bundle.paper.title.strip() or "unknown"


def _source_labels(bundles: list[PaperBundle]) -> list[str]:
    """`来源` 是汇总表里区分论文的唯一凭据，必须两两不同。

    标签取自 figures.json 里的 output_dir，而它可能为空或重名
    （render_corpus_table 拿不到 `--dirs` 的真实路径），重名时按出现顺序
    补 `#2` / `#3`，否则 groupby 会把两篇悄悄并成一篇。
    """
    labels: list[str] = []
    seen: dict[str, int] = {}
    for bundle in bundles:
        base = _bundle_source(bundle)
        seen[base] = seen.get(base, 0) + 1
        labels.append(base if seen[base] == 1 else f"{base} #{seen[base]}")
    return labels


def _paper_summary_rows(
    bundles: Iterable[PaperBundle], labels: Iterable[str]
) -> list[dict[str, Any]]:
    rows = []
    for bundle, label in zip(bundles, labels):
        panels = [p for f in bundle.figures for p in f.panels]
        rows.append(
            {
                "来源": label,
                "论文标题": _text(bundle.paper.title, ""),
                "年份": bundle.paper.year if bundle.paper.year is not None else "",
                "venue": _text(bundle.paper.venue, ""),
                "DOI": _text(bundle.paper.doi, ""),
                "domain": _val(bundle.domain),
                "图数": len(bundle.figures),
                "panel数": len(panels),
                "实验组数": sum(len(p.groups) for p in panels),
                "需复核panel数": sum(1 for p in panels if needs_review(p)),
                "源文件": bundle.source_pdf,
            }
        )
    return rows


def _frame(rows: list[dict[str, Any]], columns: list[str]) -> pd.DataFrame:
    """列序固定：即便某篇论文缺字段，跨文献汇总也要能纵向对齐。"""
    df = pd.DataFrame(rows, columns=columns)
    return df.fillna("")


# ── Excel 样式 ──────────────────────────────────────────────────────────────


def _style_sheet(ws: Worksheet, df: pd.DataFrame, freeze: str) -> None:
    ws.freeze_panes = freeze
    if len(df):
        ws.auto_filter.ref = ws.dimensions

    for col_idx, name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        values = [str(v) for v in df[name].tolist()]
        content_width = max([_dw(str(name))] + [_dw(v) for v in values]) if values else _dw(str(name))
        wrap = name in WRAP_COLUMNS or content_width > MAX_WIDTH
        ws.column_dimensions[letter].width = min(max(content_width + 2, MIN_WIDTH), MAX_WIDTH)

        header = ws.cell(row=1, column=col_idx)
        header.fill = HEADER_FILL
        header.font = HEADER_FONT
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            # 以 `=` 开头的文本会被 openpyxl 当公式写出，打开即 #NAME?
            if isinstance(cell.value, str) and cell.data_type == "f":
                cell.data_type = "s"
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
            if name == "是否需复核" and cell.value == "是":
                cell.fill = REVIEW_FILL

    ws.row_dimensions[1].height = 30


def _write_xlsx(sheets: list[tuple[str, pd.DataFrame, str]], path: Path) -> Path:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df, freeze in sheets:
            df.to_excel(writer, sheet_name=name, index=False)
            _style_sheet(writer.sheets[name], df, freeze=freeze)
    return path


def _write_csv(df: pd.DataFrame, path: Path) -> Path:
    # utf-8-sig：没有 BOM 的话 Excel 双击打开中文会是乱码
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


# ── 入口 ────────────────────────────────────────────────────────────────────


def _summary_frame(bundle: PaperBundle) -> pd.DataFrame:
    """四列速览表的 Excel 版。行结构与笔记里那张同源，避免两处各写一套。"""
    from pfn.models import review_reasons
    from pfn.render_notes import figure_summary_rows

    all_reasons = review_reasons(bundle)
    rows = []
    for fig in bundle.figures:
        reasons = all_reasons.get(fig.figure_id, [])
        for row in figure_summary_rows(fig, reasons, blank_repeats=False):
            # 行构造是给 markdown 用的，`<br>` 在 Excel 里得还原成真换行
            cells = [c.replace("<br>", "\n") for c in row]
            rows.append(dict(zip(SUMMARY_COLUMNS, [fig.figure_id, *cells])))
    return _frame(rows, SUMMARY_COLUMNS)


def render_tables(bundle: PaperBundle, out_dir: Path) -> list[Path]:
    """单篇长表，返回 [xlsx, csv]。一行 = 一个实验组。

    xlsx 含两个 sheet：`实验组长表`（全字段，跨文献筛选用）与
    `图速览`（四列，一屏看完一张图做了什么）。csv 只导长表。
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    df = _frame(_bundle_rows(bundle), COLUMNS)
    xlsx = _write_xlsx(
        [(SHEET_GROUPS, df, "A2"), (SHEET_SUMMARY, _summary_frame(bundle), "B2")],
        out_dir / f"{TABLE_BASENAME}.xlsx",
    )
    csv = _write_csv(df, out_dir / f"{TABLE_BASENAME}.csv")
    return [xlsx, csv]


def render_corpus_table(bundles: list[PaperBundle], out_dir: Path) -> Path:
    """跨文献汇总：多篇长表纵向合并，返回 xlsx 路径（同名 csv 一并写出）。

    第一列 `来源` 标出行来自哪篇，配合 Excel 自动筛选即可回答
    「所有用了 X 模型 / X 剂量的实验分别出自哪几篇、结论是否一致」。
    """
    if not bundles:
        raise ValueError("render_corpus_table 需要至少一篇 PaperBundle")

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    labels = _source_labels(bundles)
    rows: list[dict[str, Any]] = []
    for bundle, label in zip(bundles, labels):
        rows.extend(_bundle_rows(bundle, source=label))

    df = _frame(rows, CORPUS_COLUMNS)
    papers = _frame(_paper_summary_rows(bundles, labels), PAPER_COLUMNS)

    xlsx = _write_xlsx(
        [(SHEET_GROUPS, df, "B2"), (SHEET_PAPERS, papers, "A2")],
        out_dir / f"{CORPUS_BASENAME}.xlsx",
    )
    _write_csv(df, out_dir / f"{CORPUS_BASENAME}.csv")
    return xlsx
