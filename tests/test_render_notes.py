"""L4 渲染层验收测试 —— 笔记 + 长表。

样本 `fixtures/sample_bundle.json` 来自一篇真实的 Scientific Reports 论文
（3 图 / 4 panel / 13 实验组，含跨页图与 schematic 图），所有断言都对着它跑。

可直接 `python tests/test_render_notes.py` 运行，也可 `pytest tests/`。
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from pfn.models import PaperBundle  # noqa: E402
from pfn.render_notes import NOTES_FILENAME, render_notes  # noqa: E402
from pfn.render_table import COLUMNS, CORPUS_COLUMNS, render_corpus_table, render_tables  # noqa: E402

FIXTURE = ROOT / "tests" / "fixtures" / "sample_bundle.json"

EXPECTED_GROUPS = 13  # Fig1: 4+4+3, Fig2: 2, Fig6: schematic 无分组
EXPECTED_PANELS = 4


@pytest.fixture(scope="module")
def bundle() -> PaperBundle:
    return PaperBundle.model_validate_json(FIXTURE.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def notes(bundle: PaperBundle, tmp_path_factory) -> str:
    out = tmp_path_factory.mktemp("notes")
    path = render_notes(bundle, out)
    assert path.name == NOTES_FILENAME
    return path.read_text(encoding="utf-8")


# ── 笔记 ────────────────────────────────────────────────────────────────────


def test_notes_never_leaks_sentinel(notes: str) -> None:
    """铁律：`not_reported` 是内部哨兵值，绝不能出现在给人看的笔记里。"""
    assert "not_reported" not in notes
    assert "未说明" in notes  # 该缺的字段要显式说缺，不是静默省略


def test_notes_covers_every_figure_panel_group(bundle: PaperBundle, notes: str) -> None:
    for fig in bundle.figures:
        assert f"## " in notes and fig.figure_id in notes
        for panel in fig.panels:
            assert f"Panel {panel.label}" in notes
            for group in panel.groups:
                assert group.name in notes, f"{fig.figure_id}/{panel.label} 组 {group.name} 丢失"


def test_notes_marks_inferred_panels(notes: str) -> None:
    """evidence.inferred 非空的 panel 必须被 ⚠ 标出，这是人工复核入口。"""
    # Panel A（design.controlled 推断）与 Panel C（n / replicates 推断）
    assert "### ⚠ Panel A" in notes
    assert "### ⚠ Panel C" in notes
    assert "### Panel B" in notes  # inferred 为空，不该被标记
    assert "模型推断 → `design.controlled`" in notes
    assert "## 复核清单" in notes
    assert "`groups[].n`" in notes


def _section(notes: str, figure_id: str) -> str:
    """取某张图的小节。⚠ 前缀在这里剥掉，调用方一律传裸 figure_id。

    否则「⚠ Figure 1」会把不相干的用例（比如速览表那几条）也绑到 ⚠ 逻辑上，
    ⚠ 一坏它们跟着红，失败位置指向的却不是真正坏掉的东西。
    """
    parts = re.split(r"^## ", notes, flags=re.MULTILINE)
    for part in parts:
        if part.removeprefix("⚠ ").startswith(figure_id):
            return part
    raise AssertionError(f"找不到 {figure_id} 小节")


def test_schematic_has_no_empty_table(notes: str) -> None:
    """示意图没有 panels，必须单独排版，不能出现空表格骨架。"""
    section = _section(notes, "Figure 6")
    assert not any(line.startswith("|") for line in section.splitlines())
    assert "本图为示意图" in section
    assert "Panel" not in section


def test_multipage_figure_lists_all_images(notes: str) -> None:
    """跨页图（Fig 2 在 p6/p8）两张切图都要列出来。"""
    section = _section(notes, "Figure 2")
    assert "figures/fig_02_p6.png" in section
    assert "figures/fig_02_p8.png" in section
    assert "共 2 张切图" in section


def _table_blocks(notes: str) -> list[list[str]]:
    blocks, current = [], []
    for line in notes.splitlines():
        if line.startswith("|"):
            current.append(line)
        elif current:
            blocks.append(current)
            current = []
    return blocks + ([current] if current else [])


def test_notes_tables_are_wellformed(notes: str) -> None:
    """每张 markdown 表的列数必须与表头一致，否则渲染器会错行。"""
    blocks = _table_blocks(notes)
    assert len(blocks) >= 5  # 元信息 / 速览 / 3 张分组表 / 报告率
    for block in blocks:
        assert len(block) >= 3, f"表至少要有表头+分隔+1 行：{block}"
        assert set(block[1]) <= set("|- "), f"第二行必须是分隔行：{block[1]}"
        # 转义过的 `\|` 是单元格内容，不参与列数计算
        widths = {line.count("|") - line.count("\\|") for line in block}
        assert len(widths) == 1, f"表格列数不一致：{block}"


def test_clean_figure_mindmap_stays_unmarked(tmp_path: Path) -> None:
    """三路判定都干净时传空列表（= 已核对），导图不该标 ⚠。

    锁的是我方的传参语义：`review_reasons()` 查无理由传 `[]`，而不是不传
    （不传 = None = 拿不到 bundle 的降级，会退回只看自报信息）。
    """
    from pfn.models import Evidence, FigureRecord, Group, Panel

    bundle = PaperBundle(
        figures=[
            FigureRecord(
                figure_id="Figure 1",
                question="干净的图",
                confidence="high",
                panels=[
                    Panel(
                        label="A",
                        experiment="对照实验",
                        groups=[Group(name="对照", role="control")],
                        evidence=Evidence(from_caption=["experiment"]),
                    )
                ],
            )
        ]
    )
    text = render_notes(bundle, tmp_path).read_text(encoding="utf-8")
    block = text.split("```mermaid", 1)[1].split("```", 1)[0]
    assert "⚠" not in block, f"干净的图被误标：{block}"
    assert "## Figure 1" in text and "## ⚠ Figure 1" not in text


def test_notes_embeds_figure_mindmap(notes: str) -> None:
    """PLAN §5 ②：单图分组结构导图嵌在该图下方；示意图没有分组，不该嵌。"""
    assert notes.count("**分组结构导图**") == 2  # Figure 1 / Figure 2
    assert "```mermaid" in notes
    assert "mindmap" in _section(notes, "Figure 1")
    assert "分组结构导图" not in _section(notes, "Figure 6")


def _fabricated_bundle() -> PaperBundle:
    """SKILL.md 的反例：模型在 n / stats.test 上脑补，同时谎称 inferred 为空、
    自报 high。只有 merge 哨兵的 `[复核]` 警告能抓住它。"""
    from pfn.models import Evidence, FigureRecord, Group, Panel, Stats

    return PaperBundle(
        figures=[
            FigureRecord(
                figure_id="Figure 4",
                caption="Fig. 4. Keap1/Nrf2 drives HMOX1 upregulation.",
                question="DMC-GF 是否经 Keap1/Nrf2 上调 HMOX1？",
                confidence="high",
                panels=[
                    Panel(
                        label="A",
                        experiment="Western blot 检测 Nrf2 / HMOX1",
                        groups=[Group(name="DMC-GF", role="treatment", n="3")],
                        stats=Stats(test="t-test"),
                        evidence=Evidence(inferred=[]),  # ← 谎称零推断
                    )
                ],
            )
        ],
        warnings=[
            "[复核] figure_04.json · Figure 4 panels[0](A): 填了 n='3' 却没在 "
            "evidence 里溯源——n 是最高危字段，请确认不是脑补，否则改回 not_reported",
            "[复核] figure_04.json · Figure 4 panels[0](A): 填了 stats.test='t-test' "
            "却没在 evidence 里溯源——看见星号不等于知道用了什么检验",
        ],
    )


def test_sentinel_flags_fabricated_record(tmp_path: Path) -> None:
    """自报干净但被哨兵抓住的记录，必须在速览表 / 小节标题 / 复核清单三处都带 ⚠。

    回归点：早先只看 panel 自报的 `evidence.inferred`，这条造假记录反而显示
    「高置信度 + 复核栏空白」，比诚实抽取的图看着还干净。
    """
    text = render_notes(_fabricated_bundle(), tmp_path).read_text(encoding="utf-8")

    overview = [ln for ln in text.splitlines() if ln.startswith("| **Figure 4**")]
    assert len(overview) == 1, "速览表里找不到 Figure 4"
    assert "⚠" in overview[0], f"速览表未标记造假记录：{overview[0]}"
    assert "高 ⚠" in overview[0], f"自报 high 却被哨兵抓住，置信度列应示警：{overview[0]}"

    assert "## ⚠ Figure 4" in text, "图小节标题缺 ⚠"
    section = _section(text, "Figure 4")
    assert "需人工复核" in section, "图小节缺少复核提示块"

    # 断言范围收到提示块本身：整节还含嵌入的导图，那是别人渲染的，
    # 用整节判定会让本测试可能因为对方的文案而通过（或红），测错了对象
    callout = section.split("> ⚠ **需人工复核**", 1)[1].split("### ", 1)[0]
    assert "校验哨兵" in callout, "复核提示块未列出哨兵理由"
    assert "n='3'" in callout, "哨兵理由的具体内容应原样带出，便于直接核对"
    assert "Panel A" in callout, "提示块应标出理由归属哪个 panel"

    checklist = text.split("## 复核清单", 1)[1]
    assert "Figure 4" in checklist, "复核清单漏了造假记录"
    assert "哨兵单独抓出" in checklist

    # 嵌进本节的导图必须同样标记：不给 render_figure_mindmap 传 reasons 时它会
    # 退化成只看自报信息，于是「⚠ 标题 + 干净导图」自相矛盾地并排出现
    block = section.split("```mermaid", 1)[1].split("```", 1)[0]
    assert "⚠" in block, f"嵌入的导图漏标了哨兵命中的图：{block}"


def test_sentinel_marks_only_the_named_panel(tmp_path: Path) -> None:
    """多 panel 图 + 哨兵只命中一个 panel：**只有那个 panel 带 ⚠**，其余不带。

    这是 `ReviewReason.panel` 落地前做不到的——旧实现在 `w.split(":", 1)` 那步
    把 `panels[2](C)` 切掉了，5 个 panel 只能一起标或一起不标。
    """
    from pfn.models import FigureRecord, Group, Panel

    bundle = PaperBundle(
        figures=[
            FigureRecord(
                figure_id="Figure 3",
                confidence="high",
                panels=[
                    Panel(
                        label=label,
                        experiment=f"panel {label} 的实验",
                        groups=[Group(name=f"g{label}", role="treatment", n="3")],
                    )
                    for label in "ABCDE"
                ],
            )
        ],
        warnings=[
            "[复核] figure_03.json · Figure 3 panels[2](C): 填了 n='3' 却没在 "
            "evidence 里溯源——n 是最高危字段，请确认不是脑补"
        ],
    )
    text = render_notes(bundle, tmp_path).read_text(encoding="utf-8")

    headings = [ln for ln in text.splitlines() if ln.startswith("### ")]
    assert len(headings) == 5
    marked = [ln for ln in headings if "⚠" in ln]
    assert len(marked) == 1, f"应只标 Panel C，实得 {marked}"
    assert marked[0].startswith("### ⚠ Panel C"), marked[0]

    # 只取 panel 区（到「整图结论」为止），否则最后一个 panel 会串到文末复核清单
    panels_only = text.split("**整图结论**", 1)[0]
    sections = {
        blk.split("\n", 1)[0].lstrip("⚠ ").strip(): blk
        for blk in panels_only.split("### ")[1:]
    }
    assert set(sections) >= {f"Panel {c} · panel {c} 的实验" for c in "ABCDE"}

    # 哨兵理由要落在 Panel C 自己的小节里，而不是只挂在图级
    panel_c = sections["Panel C · panel C 的实验"]
    assert "校验哨兵" in panel_c and "n='3'" in panel_c
    for other in ("A", "B", "D", "E"):
        assert "校验哨兵" not in sections[f"Panel {other} · panel {other} 的实验"], (
            f"Panel {other} 不该出现哨兵理由"
        )

    # 图级提示块要标出归属，读者不用逐个 panel 找
    assert "Panel C · **校验哨兵**" in text


def test_review_marks_agree_across_sections(bundle: PaperBundle, notes: str) -> None:
    """速览表打 ⚠ 的图，正文小节与复核清单必须同样打 ⚠，三处口径不能打架。"""
    from pfn.models import review_reasons

    flagged = set(review_reasons(bundle))
    checklist = notes.split("## 复核清单", 1)[1]
    for fig in bundle.figures:
        row = next(ln for ln in notes.splitlines() if ln.startswith(f"| **{fig.figure_id}**"))
        marked = "⚠" in row
        assert marked == (fig.figure_id in flagged), f"{fig.figure_id} 速览表标记不一致"
        assert (f"## ⚠ {fig.figure_id}" in notes) == marked, f"{fig.figure_id} 小节标题不一致"
        assert (fig.figure_id in checklist) == marked, f"{fig.figure_id} 复核清单不一致"


# ── 四列速览表 ──────────────────────────────────────────────────────────────


def test_summary_table_embedded_before_panels(notes: str) -> None:
    """速览表要在 panel 详细小节**之前**——先速览再下钻。"""
    section = _section(notes, "Figure 1")
    assert "**速览**" in section
    # 用裸 `### ` 判位置，别把顺序断言绑到 ⚠ 上——⚠ 归 review_reasons 的用例管
    assert section.index("**速览**") < section.index("### ")
    header = "| 小图 | 实验方法 | 分组 | 结果 |"
    assert header in section


def test_summary_table_blanks_repeated_cells(notes: str) -> None:
    """同一 panel 的后续行留空前两列；⚠ 打在「小图」列。"""
    section = _section(notes, "Figure 1")
    rows = [ln for ln in section.splitlines() if ln.startswith("|")]
    body = [r for r in rows if not set(r) <= set("|- ") and "小图" not in r]
    assert body[0].startswith("| ⚠ A |"), body[0]
    # Panel A 有 4 组，后 3 行前两列必须为空，否则同一实验会被读成四个实验
    assert [r.split("|")[1].strip() for r in body[1:4]] == ["", "", ""]
    assert body[4].startswith("| B |"), body[4]  # 下一个 panel 重新出现标号
    assert body[8].startswith("| ⚠ C |"), body[8]


def test_summary_table_drops_redundant_condition(notes: str) -> None:
    """组名已含剂量时不再重复处理条件，否则四行里同一句话写四遍。"""
    section = _section(notes, "Figure 1")
    assert "| ● U251 · 12 h |" in section
    assert "DMC-GF 0-4.5 μM 梯度，处理 12 h" not in section.split("### ")[0]
    # 组名说不清的照旧带上条件
    assert "○ 0 μM（溶剂对照，不加 DMC-GF）" in section


def test_summary_merges_significance_only(notes: str) -> None:
    """带统计标记的 vs_control 并进结果列；描述性对比留在详细表里。"""
    section = _section(notes, "Figure 1")
    summary = section.split("### ")[0]
    assert "约 75%（*** p<0.001）" in summary
    assert "相对 0 μM 对照下降约 45%" not in summary


def test_figure_tables_file(bundle: PaperBundle, tmp_path: Path) -> None:
    from pfn.render_notes import TABLES_FILENAME, render_figure_tables

    path = render_figure_tables(bundle, tmp_path)
    assert path.name == TABLES_FILENAME
    text = path.read_text(encoding="utf-8")
    assert "not_reported" not in text
    assert text.count("| 小图 | 实验方法 | 分组 | 结果 |") == 2  # Fig 1 / Fig 2
    # schematic 不出空表
    assert "本图为示意图，不含实验分组。" in text
    section6 = text.split("## Figure 6", 1)[1]
    assert not any(ln.startswith("|") for ln in section6.splitlines())


def test_render_notes_also_writes_tables(bundle: PaperBundle, tmp_path: Path) -> None:
    """CLI 只收 render_notes 的返回值，速览表必须由它顺带写出，否则不会产出。"""
    from pfn.render_notes import TABLES_FILENAME

    render_notes(bundle, tmp_path)
    assert (tmp_path / TABLES_FILENAME).is_file()


def test_summary_escapes_pipe_in_names(tmp_path: Path) -> None:
    """组名里出现 `|` 不算罕见，不转义会把表结构冲散。"""
    from pfn.models import FigureRecord, Group, Panel

    bundle = PaperBundle(
        figures=[
            FigureRecord(
                figure_id="Figure 1",
                panels=[
                    Panel(
                        label="A",
                        experiment="A|B 对比",
                        groups=[Group(name="模型 A|B", role="treatment", result="胜率 60%")],
                    )
                ],
            )
        ]
    )
    text = render_notes(bundle, tmp_path).read_text(encoding="utf-8")
    row = next(ln for ln in text.splitlines() if "模型 A" in ln and ln.startswith("|"))
    assert r"模型 A\|B" in row
    assert row.count("|") - row.count(r"\|") == 5  # 四列 = 五根竖线


def test_hostile_caption_cannot_break_markdown(tmp_path: Path) -> None:
    """图题/组名是 PDF 与模型来的文本，直接写进正文属于未经转义的路径。

    ``` 会把围栏配对搞乱（笔记里内嵌了 mermaid 块，错位会毁掉整个文件），
    `<script>` 会被 Obsidian 当裸 HTML 渲染。
    """
    from pfn.models import FigureRecord, Group, Panel

    hostile = "Fig. 1. Bad ```mermaid\nfake``` and <script>alert(1)</script> and &lt;b&gt;"
    bundle = PaperBundle(
        paper={"title": "Title with <img src=x> and `code`"},
        figures=[
            FigureRecord(
                figure_id="Figure 1",
                caption=hostile,
                panels=[
                    Panel(
                        label="A",
                        experiment="正常实验",
                        groups=[Group(name="<b>组</b>", role="treatment", result="p<0.001 且 x>1")],
                    )
                ],
            )
        ],
    )
    text = render_notes(bundle, tmp_path).read_text(encoding="utf-8")

    # 围栏必须仍然成对：只有我自己嵌的那个 mermaid 块
    assert text.count("```") % 2 == 0
    assert text.count("```") == 2, "图题里的 ``` 泄漏成了额外围栏"
    # 只中和标签起始的 `<`，落单的 `>` 开不了标签，留着原文更好读
    assert "<script>" not in text and "&lt;script>" in text
    assert "<img" not in text and "<b>" not in text
    # 已是实体的再转一层。**这不是安全措施**——markdown-it 与 python-markdown 实测
    # 都把 `&lt;script&gt;` 原样输出为转义文本，不会变成活标签。理由是保真：
    # 原文 PDF 里写的是 `&lt;b&gt;` 这 8 个字符，渲染出来就该是这 8 个字符，
    # 不转的话读者看到的是 `<b>`，与原文不符。
    assert "&amp;lt;b&amp;gt;" in text

    # 比较符不能被误伤——`p<0.001` 在生物医学图题里遍地都是
    assert "p<0.001 且 x>1" in text

    tables = (tmp_path / "figure-tables.md").read_text(encoding="utf-8")
    assert tables.count("```") == 0 and "<script>" not in tables


def test_real_caption_survives_verbatim(tmp_path: Path) -> None:
    """图题要能拿去原文 PDF 里 Ctrl+F —— 改一个字符就搜不到了。

    `log2FC > 1`、`p<0.05`、`Nrf2 & HMOX1` 在生物医学图题里是常态，
    转义规则必须放过它们。
    """
    from pfn.models import FigureRecord, Group, Panel

    caption = "Fig. 3. Genes with log2FC > 1 and p<0.05 in Nrf2 & HMOX1 pathway (n = 3, mean ± SD)."
    bundle = PaperBundle(
        figures=[
            FigureRecord(
                figure_id="Figure 3",
                caption=caption,
                panels=[Panel(label="A", experiment="x", groups=[Group(name="g", role="treatment")])],
            )
        ]
    )
    text = render_notes(bundle, tmp_path).read_text(encoding="utf-8")
    quoted = next(ln for ln in text.splitlines() if ln.startswith("> Fig. 3."))
    assert quoted[2:] == caption, f"图题被改动，Ctrl+F 会失效：{quoted[2:]}"


def test_summary_sheet_in_xlsx(bundle: PaperBundle, tmp_path: Path) -> None:
    xlsx, _ = render_tables(bundle, tmp_path)
    df = pd.read_excel(xlsx, sheet_name="图速览")
    assert list(df.columns) == ["figure_id", "小图", "实验方法", "分组", "结果"]
    assert len(df) == EXPECTED_GROUPS
    # Excel 版不留空重复值，否则一排序就和上面的行脱钩
    assert df["小图"].notna().all() and (df["小图"].astype(str).str.strip() != "").all()
    assert not df.to_string().count("<br>"), "`<br>` 是 markdown 写法，Excel 里应是真换行"


def test_notes_is_deterministic(bundle: PaperBundle, tmp_path: Path) -> None:
    """纯函数：同样输入两次渲染必须逐字节一致（无时间戳等噪声）。"""
    a = render_notes(bundle, tmp_path / "a").read_bytes()
    b = render_notes(bundle, tmp_path / "b").read_bytes()
    assert a == b


def test_notes_written_as_utf8_lf(bundle: PaperBundle, tmp_path: Path) -> None:
    """Windows 默认 GBK + CRLF；与其余 L4 产物统一为 UTF-8 + LF。"""
    raw = render_notes(bundle, tmp_path).read_bytes()
    assert b"\r\n" not in raw
    assert "凋亡率" in raw.decode("utf-8")


def test_notes_handles_empty_bundle(tmp_path: Path) -> None:
    path = render_notes(PaperBundle(), tmp_path)
    text = path.read_text(encoding="utf-8")
    assert "未提取到任何图" in text
    assert "not_reported" not in text


# ── 长表 ────────────────────────────────────────────────────────────────────


def test_tables_row_per_group(bundle: PaperBundle, tmp_path: Path) -> None:
    xlsx, csv = render_tables(bundle, tmp_path)
    assert xlsx.suffix == ".xlsx" and csv.suffix == ".csv"

    df = pd.read_excel(xlsx, sheet_name="实验组长表", dtype=str)
    assert len(df) == EXPECTED_GROUPS, f"行数应等于实验组数，实得 {len(df)}"
    assert list(df.columns) == COLUMNS
    assert "not_reported" not in df.to_string()

    # panel 级字段冗余到每一行，Excel 一次筛选即可选出同一 panel 的所有组
    assert df[df["figure_id"] == "Figure 1"].shape[0] == 11
    assert set(df["panel"]) == {"A", "B", "C"}
    # Fig1 Panel A 4 组 + Fig1 Panel C 3 组 + Fig2 Panel A 2 组
    assert df["是否需复核"].tolist().count("是") == 9


def test_csv_has_bom_for_excel(bundle: PaperBundle, tmp_path: Path) -> None:
    """没有 BOM 的话 Excel 双击打开中文是乱码。"""
    _, csv = render_tables(bundle, tmp_path)
    assert csv.read_bytes().startswith(b"\xef\xbb\xbf")
    df = pd.read_csv(csv, encoding="utf-8-sig", dtype=str)
    assert len(df) == EXPECTED_GROUPS
    assert df.loc[0, "论文标题"].startswith("The curcumin derivative")


def test_xlsx_is_styled(bundle: PaperBundle, tmp_path: Path) -> None:
    from openpyxl import load_workbook

    xlsx, _ = render_tables(bundle, tmp_path)
    ws = load_workbook(xlsx)["实验组长表"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws.column_dimensions["A"].width > 8
    assert ws.cell(row=1, column=1).font.bold
    assert ws.cell(row=2, column=1).alignment.wrap_text


def test_corpus_merges_papers(bundle: PaperBundle, tmp_path: Path) -> None:
    second = bundle.model_copy(deep=True)
    second.paper.title = "Another paper"
    second.output_dir = "notes_other"

    xlsx = render_corpus_table([bundle, second], tmp_path)
    df = pd.read_excel(xlsx, sheet_name="实验组长表", dtype=str)
    assert len(df) == EXPECTED_GROUPS * 2
    assert list(df.columns) == CORPUS_COLUMNS
    assert set(df["来源"]) == {"notes_scirep_dmcgf", "notes_other"}

    papers = pd.read_excel(xlsx, sheet_name="论文清单")
    assert len(papers) == 2
    assert papers["实验组数"].tolist() == [EXPECTED_GROUPS, EXPECTED_GROUPS]
    assert papers["需复核panel数"].tolist() == [3, 3]
    assert (xlsx.parent / "all_groups.csv").read_bytes().startswith(b"\xef\xbb\xbf")


def test_table_survives_zero_group_paper(tmp_path: Path) -> None:
    """整篇只有示意图时应写出「只有表头」的空表，而不是崩掉或造假行。"""
    from pfn.models import FigureRecord

    bundle = PaperBundle(figures=[FigureRecord(figure_id="Figure 1", figure_kind="schematic")])
    xlsx, csv = render_tables(bundle, tmp_path)
    df = pd.read_excel(xlsx)
    assert len(df) == 0
    assert list(df.columns) == COLUMNS
    assert csv.read_bytes().startswith(b"\xef\xbb\xbf")


def test_corpus_disambiguates_same_source(bundle: PaperBundle, tmp_path: Path) -> None:
    """两篇 output_dir 撞名时，来源列仍必须可区分，否则汇总会把它们并成一篇。"""
    xlsx = render_corpus_table([bundle, bundle.model_copy(deep=True)], tmp_path)
    df = pd.read_excel(xlsx, sheet_name="实验组长表", dtype=str)
    assert len(set(df["来源"])) == 2
    assert df.groupby("来源").size().tolist() == [EXPECTED_GROUPS, EXPECTED_GROUPS]


def test_corpus_rejects_empty(tmp_path: Path) -> None:
    with pytest.raises(ValueError):
        render_corpus_table([], tmp_path)


def test_group_count_matches_fixture() -> None:
    """防止 fixture 被改动后测试的期望值悄悄失真。"""
    data = json.loads(FIXTURE.read_text(encoding="utf-8"))
    panels = [p for f in data["figures"] for p in f["panels"]]
    assert len(panels) == EXPECTED_PANELS
    assert sum(len(p["groups"]) for p in panels) == EXPECTED_GROUPS


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
